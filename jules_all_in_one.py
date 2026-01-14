#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
JULES — ALL-IN-ONE (IMAP + SMTP + Excel + Telegram)

Qué hace (en 1 solo script):
1) Lee tu Excel (pipeline de contactos).
2) Detecta quién está "Pending reply" / "Contacted" sin respuesta tras X horas -> reenvía el follow-up con la oferta.
3) Lee respuestas reales del buzón (IMAP) -> escribe en el Excel:
   - Reply? (YES/NO)
   - Reply Date
   - Reply From
   - Reply Subject
   - Reply Text
   - Reply Message-ID
   y actualiza Status/Next Action con lógica simple.
4) Te manda un resumen por Telegram (@abvet_deploy_bot vía token).

Requisitos:
  pip install openpyxl python-dotenv

ENV (todo en secrets/env):
  EXCEL_PATH=/ruta/Startup_Follow_Up_Sheet_TRYONYOU.xlsx
  SHEET_NAME=Sheet1                       # opcional

  IMAP_HOST=imap.porkbun.com
  IMAP_PORT=993
  IMAP_USER=contact@tryonyou.app
  IMAP_PASS=*****

  SMTP_HOST=smtp.porkbun.com
  SMTP_PORT=587
  SMTP_USER=contact@tryonyou.app
  SMTP_PASS=*****
  FROM_EMAIL=contact@tryonyou.app         # opcional
  FROM_NAME=TRYONYOU                      # opcional

  TELEGRAM_BOT_TOKEN=*****
  TELEGRAM_CHAT_ID=*****

  FOLLOWUP_AFTER_HOURS=48                 # default 48
  DRY_RUN=1                               # 1 para probar sin enviar, 0 para enviar
  IMAP_SINCE_DAYS=7                       # default 7 (cuántos días atrás leer respuestas)
  TARGET_STATUSES=Pending reply,Contacted # default

Columnas esperadas (si faltan, el script las crea):
  Email, Name, Company (opcional), Status, Last Action Date, Next Action, Offer Sent, Notes / Comments
  Reply?, Reply Date, Reply From, Reply Subject, Reply Text, Reply Message-ID

Uso:
  python jules_all_in_one.py
"""

import os
import re
import ssl
import smtplib
import imaplib
import email
from email.header import decode_header, make_header
from email.message import EmailMessage
from datetime import datetime, timedelta, timezone
from typing import Optional, List, Dict, Tuple, Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

DATE_FMT = "%Y-%m-%d"
DT_FMT = "%Y-%m-%d %H:%M:%S %z"

STATUSES_ALLOWED = {
    "Contacted",
    "Pending reply",
    "Follow-up sent",
    "Interested",
    "Not interested",
    "Meeting scheduled",
    "Closed / In discussion",
}

DEFAULT_COLS = [
    "Email",
    "Name",
    "Company",
    "Status",
    "Last Action Date",
    "Next Action",
    "Offer Sent",
    "Notes / Comments",
    "Reply?",
    "Reply Date",
    "Reply From",
    "Reply Subject",
    "Reply Text",
    "Reply Message-ID",
]

KEYWORDS_NOT_INTERESTED = [
    "not interested", "no gracias", "no, gracias", "pass", "we'll pass", "not a fit", "not fit",
    "no encaja", "no me interesa", "no interesa", "rechazo"
]
KEYWORDS_INTERESTED = [
    "interested", "let's talk", "call", "meeting", "meet", "discuss", "deck", "nda", "dataroom",
    "me interesa", "hablemos", "llamada", "reunión", "reunion", "info", "más info", "mas info"
]

MAX_REPLY_TEXT_LENGTH = 1000  # Maximum characters to store from reply text

def env(name: str, default: Optional[str] = None) -> str:
    v = os.getenv(name, default)
    if v is None or str(v).strip() == "":
        raise SystemExit(f"Missing env var: {name}")
    return str(v).strip()

def opt_env(name: str, default: str = "") -> str:
    v = os.getenv(name, default)
    return (v or default).strip()

def now_utc() -> datetime:
    return datetime.now(timezone.utc)

def today_str() -> str:
    return now_utc().strftime(DATE_FMT)

def parse_date_safe(s: str) -> Optional[datetime]:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in (DT_FMT, DATE_FMT):
        try:
            dt = datetime.strptime(s, fmt)
            if fmt == DATE_FMT:
                dt = dt.replace(tzinfo=timezone.utc)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt
        except Exception:
            continue
    return None

def normalize_email(addr: str) -> str:
    addr = (addr or "").strip().lower()
    m = re.search(r"[\w\.\+\-]+@[\w\.\-]+\.\w+", addr)
    return m.group(0).lower() if m else ""

def decode_mime(s: str) -> str:
    if not s:
        return ""
    try:
        return str(make_header(decode_header(s)))
    except Exception:
        return s

def get_email_body(msg: email.message.Message) -> str:
    # Prefer text/plain; fallback to stripped html
    if msg.is_multipart():
        parts = msg.walk()
    else:
        parts = [msg]

    text_parts = []
    html_parts = []
    for part in parts:
        ctype = part.get_content_type()
        disp = str(part.get("Content-Disposition", "")).lower()
        if "attachment" in disp:
            continue
        try:
            payload = part.get_payload(decode=True)
        except Exception:
            payload = None
        if payload is None:
            continue
        charset = part.get_content_charset() or "utf-8"
        try:
            decoded = payload.decode(charset, errors="replace")
        except Exception:
            decoded = payload.decode("utf-8", errors="replace")

        if ctype == "text/plain":
            text_parts.append(decoded)
        elif ctype == "text/html":
            html_parts.append(decoded)

    if text_parts:
        return clean_reply_text("\n\n".join(text_parts))

    if html_parts:
        html = "\n\n".join(html_parts)
        # very light html strip
        html = re.sub(r"<(script|style)[^>]*>.*?</\1>", " ", html, flags=re.I | re.S)
        html = re.sub(r"<[^>]+>", " ", html)
        html = re.sub(r"&nbsp;", " ", html, flags=re.I)
        html = re.sub(r"&[a-z]+;", " ", html, flags=re.I)
        return clean_reply_text(html)

    return ""

def clean_reply_text(txt: str) -> str:
    # remove quoted text
    lines = txt.split("\n")
    out = []
    for line in lines:
        if line.strip().startswith(">"):
            continue
        if re.match(r"^On .* wrote:$", line.strip()):
            break
        out.append(line)
    txt = "\n".join(out).strip()
    txt = re.sub(r"\n{3,}", "\n\n", txt)
    return txt[:MAX_REPLY_TEXT_LENGTH]  # limit size

def send_telegram(token: str, chat_id: str, text: str) -> None:
    """Send a message to Telegram using bot API."""
    import urllib.request
    import json
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    data = json.dumps({"chat_id": chat_id, "text": text, "parse_mode": "Markdown"}).encode()
    req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            resp.read()
    except Exception as e:
        print(f"[WARN] Telegram failed: {e}")

def ensure_columns(ws: Worksheet, cols: List[str]) -> None:
    """Ensure all required columns exist in the worksheet header row."""
    header_row = ws[1]
    existing = [str(c.value or "").strip() for c in header_row]
    for col in cols:
        if col not in existing:
            # add to next available column
            next_col = len(existing) + 1
            ws.cell(1, next_col, col)
            existing.append(col)

def get_col_index(ws: Worksheet, col_name: str) -> Optional[int]:
    """Get the column index for a given column name."""
    for idx, cell in enumerate(ws[1], start=1):
        if str(cell.value or "").strip() == col_name:
            return idx
    return None

def load_excel_data(path: str, sheet_name: str = "Sheet1") -> Tuple[Workbook, Worksheet, Dict[str, int], List[Dict[str, Any]]]:
    """Load Excel data and return workbook, worksheet, column mapping, and row data."""
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    ensure_columns(ws, DEFAULT_COLS)
    
    # build map
    col_map = {}
    for idx, cell in enumerate(ws[1], start=1):
        col_map[str(cell.value or "").strip()] = idx
    
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        data = {}
        for col_name, col_idx in col_map.items():
            val = row[col_idx - 1].value
            data[col_name] = str(val).strip() if val else ""
        if data.get("Email"):
            data["_row_num"] = row[0].row
            rows.append(data)
    
    return wb, ws, col_map, rows

def save_excel(wb: Workbook, path: str) -> None:
    """Save the workbook to the specified path."""
    wb.save(path)

def send_followup_email(smtp_host: str, smtp_port: int, smtp_user: str, smtp_pass: str, 
                       from_email: str, from_name: str, to_email: str, to_name: str, company: str,
                       patent_number: str = "PCT/EP2025/067317", pricing: str = "€4,900/month") -> None:
    """
    Send a follow-up email to a contact.
    
    Args:
        smtp_host: SMTP server hostname
        smtp_port: SMTP server port
        smtp_user: SMTP username
        smtp_pass: SMTP password
        from_email: Sender email address
        from_name: Sender name
        to_email: Recipient email address
        to_name: Recipient name
        company: Recipient company name
        patent_number: Patent number to include in email (default: PCT/EP2025/067317)
        pricing: Pricing information to include (default: €4,900/month)
    """
    msg = EmailMessage()
    msg["Subject"] = "Follow-up: TRYONYOU Exclusive Pilot Opportunity"
    msg["From"] = f"{from_name} <{from_email}>"
    msg["To"] = to_email
    
    body = f"""Hi {to_name},

I wanted to follow up on my previous message regarding TRYONYOU.

We're offering an exclusive pilot program for {company or 'your company'} to eliminate returns with our AI-powered virtual try-on technology.

✅ Reduce returns to 0%
✅ Patent-protected technology ({patent_number})
✅ Proven results with Galeries Lafayette & Hub71
✅ {pricing} pilot pricing

Would you be interested in a quick call to discuss how we can help?

Best regards,
{from_name}
{from_email}
"""
    
    msg.set_content(body)
    
    context = ssl.create_default_context()
    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.starttls(context=context)
        server.login(smtp_user, smtp_pass)
        server.send_message(msg)

def fetch_imap_replies(imap_host: str, imap_port: int, imap_user: str, imap_pass: str, since_days: int = 7) -> List[Dict[str, Any]]:
    """
    Fetch email replies from IMAP server.
    
    Args:
        imap_host: IMAP server hostname
        imap_port: IMAP server port
        imap_user: IMAP username
        imap_pass: IMAP password
        since_days: Number of days to look back for emails (default: 7)
    
    Returns:
        List of reply dictionaries with keys: from, subject, date, body, message_id
    """
    context = ssl.create_default_context()
    mail = imaplib.IMAP4_SSL(imap_host, imap_port, ssl_context=context)
    mail.login(imap_user, imap_pass)
    mail.select("INBOX")
    
    since_date = (now_utc() - timedelta(days=since_days)).strftime("%d-%b-%Y")
    _, msg_nums = mail.search(None, f'(SINCE {since_date})')
    
    replies = []
    for num in msg_nums[0].split():
        _, msg_data = mail.fetch(num, "(RFC822)")
        raw = msg_data[0][1]
        msg = email.message_from_bytes(raw)
        
        from_addr = normalize_email(msg.get("From", ""))
        subject = decode_mime(msg.get("Subject", ""))
        date_str = msg.get("Date", "")
        msg_id = msg.get("Message-ID", "")
        body = get_email_body(msg)
        
        # parse date
        try:
            dt = email.utils.parsedate_to_datetime(date_str)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
        except Exception:
            dt = now_utc()
        
        replies.append({
            "from": from_addr,
            "subject": subject,
            "date": dt,
            "body": body,
            "message_id": msg_id,
        })
    
    mail.close()
    mail.logout()
    return replies

def classify_interest(text: str) -> str:
    text_lower = text.lower()
    for kw in KEYWORDS_NOT_INTERESTED:
        if kw in text_lower:
            return "Not interested"
    for kw in KEYWORDS_INTERESTED:
        if kw in text_lower:
            return "Interested"
    return "Pending reply"

def main():
    print("🤖 JULES ALL-IN-ONE starting...")
    
    # Load config
    excel_path = env("EXCEL_PATH")
    sheet_name = opt_env("SHEET_NAME", "Sheet1")
    
    imap_host = env("IMAP_HOST")
    imap_port = int(opt_env("IMAP_PORT", "993"))
    imap_user = env("IMAP_USER")
    imap_pass = env("IMAP_PASS")
    
    smtp_host = env("SMTP_HOST")
    smtp_port = int(opt_env("SMTP_PORT", "587"))
    smtp_user = env("SMTP_USER")
    smtp_pass = env("SMTP_PASS")
    from_email = opt_env("FROM_EMAIL", smtp_user)
    from_name = opt_env("FROM_NAME", "TRYONYOU")
    
    patent_number = opt_env("PATENT_NUMBER", "PCT/EP2025/067317")
    pricing = opt_env("PRICING", "€4,900/month")
    
    telegram_token = opt_env("TELEGRAM_BOT_TOKEN")
    telegram_chat = opt_env("TELEGRAM_CHAT_ID")
    
    followup_hours = int(opt_env("FOLLOWUP_AFTER_HOURS", "48"))
    dry_run = int(opt_env("DRY_RUN", "1")) == 1
    imap_since_days = int(opt_env("IMAP_SINCE_DAYS", "7"))
    target_statuses = [s.strip() for s in opt_env("TARGET_STATUSES", "Pending reply,Contacted").split(",")]
    
    print(f"📊 Excel: {excel_path}")
    print(f"📧 IMAP: {imap_user}@{imap_host}")
    print(f"📤 SMTP: {smtp_user}@{smtp_host}")
    print(f"🔔 Telegram: {'✅' if telegram_token else '❌'}")
    print(f"🏃 Dry run: {dry_run}")
    print()
    
    # Load Excel
    wb, ws, col_map, rows = load_excel_data(excel_path, sheet_name)
    print(f"✅ Loaded {len(rows)} contacts from Excel")
    
    # Fetch IMAP replies
    print(f"📬 Fetching IMAP replies (last {imap_since_days} days)...")
    replies = fetch_imap_replies(imap_host, imap_port, imap_user, imap_pass, imap_since_days)
    print(f"✅ Found {len(replies)} emails")
    
    # Build reply map by email
    reply_map = {}
    for r in replies:
        reply_map[r["from"]] = r
    
    # Process contacts
    followups_sent = 0
    replies_processed = 0
    
    for row_data in rows:
        email_addr = normalize_email(row_data.get("Email", ""))
        if not email_addr:
            continue
        
        status = row_data.get("Status", "").strip()
        last_action = row_data.get("Last Action Date", "").strip()
        reply_status = row_data.get("Reply?", "").strip()
        
        row_num = row_data["_row_num"]
        
        # Check for reply
        if email_addr in reply_map and reply_status != "YES":
            reply = reply_map[email_addr]
            print(f"  💬 Reply from {email_addr}: {reply['subject']}")
            
            # Update Excel
            ws.cell(row_num, col_map["Reply?"], "YES")
            ws.cell(row_num, col_map["Reply Date"], reply["date"].strftime(DT_FMT))
            ws.cell(row_num, col_map["Reply From"], reply["from"])
            ws.cell(row_num, col_map["Reply Subject"], reply["subject"])
            ws.cell(row_num, col_map["Reply Text"], reply["body"])
            ws.cell(row_num, col_map["Reply Message-ID"], reply["message_id"])
            
            # Classify interest
            new_status = classify_interest(reply["body"])
            ws.cell(row_num, col_map["Status"], new_status)
            
            if new_status == "Interested":
                ws.cell(row_num, col_map["Next Action"], "Schedule meeting")
            elif new_status == "Not interested":
                ws.cell(row_num, col_map["Next Action"], "Close")
            else:
                ws.cell(row_num, col_map["Next Action"], "Follow up")
            
            replies_processed += 1
        
        # Check for follow-up
        if status in target_statuses and reply_status != "YES":
            last_dt = parse_date_safe(last_action)
            if last_dt:
                hours_since = (now_utc() - last_dt).total_seconds() / 3600
                if hours_since >= followup_hours:
                    name = row_data.get("Name", "").strip()
                    company = row_data.get("Company", "").strip()
                    
                    print(f"  📤 Follow-up needed for {email_addr} ({hours_since:.1f}h since last action)")
                    
                    if not dry_run:
                        try:
                            send_followup_email(
                                smtp_host, smtp_port, smtp_user, smtp_pass,
                                from_email, from_name, email_addr, name, company,
                                patent_number, pricing
                            )
                            ws.cell(row_num, col_map["Status"], "Follow-up sent")
                            ws.cell(row_num, col_map["Last Action Date"], now_utc().strftime(DT_FMT))
                            ws.cell(row_num, col_map["Next Action"], "Wait for reply")
                            followups_sent += 1
                            print(f"    ✅ Sent")
                        except Exception as e:
                            print(f"    ❌ Failed: {e}")
                    else:
                        print(f"    🏃 DRY RUN - would send")
                        followups_sent += 1
    
    # Save Excel
    save_excel(wb, excel_path)
    print(f"\n💾 Saved Excel: {excel_path}")
    
    # Summary
    summary = f"""🤖 *JULES ALL-IN-ONE Summary*

📊 Contacts processed: {len(rows)}
📬 Replies found: {len(replies)}
💬 Replies processed: {replies_processed}
📤 Follow-ups sent: {followups_sent}

🏃 Dry run: {dry_run}
"""
    
    print()
    print(summary)
    
    if telegram_token and telegram_chat:
        send_telegram(telegram_token, telegram_chat, summary)
        print("✅ Telegram notification sent")
    
    print("\n✨ JULES completed!")

if __name__ == "__main__":
    main()
